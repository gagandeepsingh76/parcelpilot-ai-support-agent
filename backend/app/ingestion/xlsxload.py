"""Load ParcelPilot_Assessment_Data.xlsx sheets into plain dicts.

Tolerances built in for the real pack vs our fixture:
- sheet names matched case-insensitively by keyword;
- headers normalised (case/spacing/punctuation insensitive);
- date/datetime cells converted to UTC ISO-8601 strings with Z suffix;
- README sheet accepts two-column key/value rows or 'key=value'/'key: value'.
"""

from __future__ import annotations

import datetime as _dt
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook


class IngestionError(RuntimeError):
    """User-facing ingestion failure with guidance."""


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def _cell_to_value(value):
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=_dt.timezone.utc)
        return value.astimezone(_dt.timezone.utc).isoformat().replace("+00:00", "Z")
    if isinstance(value, _dt.date):
        return value.isoformat() + "T00:00:00Z"
    if isinstance(value, _dt.time):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def open_workbook(path: str | Path):
    try:
        return load_workbook(str(path), data_only=True, read_only=True)
    except zipfile.BadZipFile as exc:
        raise IngestionError(
            f"'{Path(path).name}' is not a valid .xlsx workbook "
            "(legacy .xls must be converted first)"
        ) from exc


def find_sheet(wb, keyword: str):
    for name in wb.sheetnames:
        if keyword.lower() in str(name).strip().lower():
            return wb[name]
    raise IngestionError(
        f"workbook has no sheet matching '{keyword}' (found: {', '.join(wb.sheetnames)})"
    )


def read_key_values(ws) -> dict[str, str]:
    """README-style sheet -> dict. Supports A/B columns or inline key=value / key: value."""
    meta: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        first = str(cells[0]).strip()
        if len(cells) >= 2:
            key, val = first, cells[1]
        else:
            for sep in ("=", ":"):
                if sep in first:
                    key, val = first.split(sep, 1)[0].strip(), first.split(sep, 1)[1].strip()
                    break
            else:
                continue
        if val is None or str(val).strip() == "":
            continue
        meta[_norm_header(key)] = str(val).strip()
    return meta


def require_snapshot(meta: dict[str, str]) -> str:
    for key, value in meta.items():
        if "snapshot" in key and "utc" in key:
            return value
    for key, value in meta.items():
        if "snapshot" in key:
            return value
    raise IngestionError(
        "workbook README sheet does not declare a dataset snapshot timestamp "
        "(looked for a key containing 'snapshot'); every time-based calculation "
        "depends on it"
    )


def load_table(ws, required: set[str], label: str) -> list[dict]:
    """Generic row loader: validates required columns, returns list of dicts."""
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise IngestionError(f"'{label}' sheet is empty")
    headers = [_norm_header(h) if h is not None else "" for h in header_row]
    missing = [c for c in sorted(required) if c not in headers]
    if missing:
        raise IngestionError(
            f"'{label}' sheet is missing required column(s): {', '.join(missing)} "
            f"(found: {', '.join(h for h in headers if h)})"
        )
    records: list[dict] = []
    for idx, raw in enumerate(rows_iter, start=2):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        record = {}
        for header, value in zip(headers, raw):
            if header:
                record[header] = _cell_to_value(value)
        record["_row_number"] = idx
        records.append(record)
    if not records:
        raise IngestionError(f"'{label}' sheet has a header but no data rows")
    return records


def load_accounts(wb) -> list[dict]:
    return load_table(find_sheet(wb, "account"), {"account_id"}, "accounts")


def load_orders(wb) -> list[dict]:
    return load_table(find_sheet(wb, "order"), {"order_id", "account_id"}, "orders")


def load_tickets(wb) -> list[dict]:
    return load_table(find_sheet(wb, "ticket"), {"ticket_id", "account_id"}, "tickets")
