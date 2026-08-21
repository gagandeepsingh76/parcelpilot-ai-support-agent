"""Generate a SYNTHETIC stand-in data pack for development and tests.

The real candidate pack is proprietary and was not available on disk when this
repo was scaffolded (ASSUMPTIONS.md #1). This script produces fixtures matching
the documented pack layout (data_pack/README.md) so every downstream step can
be built and verified end to end. When the real pack arrives, drop it into
data_pack/ and re-run ingestion -- no code changes needed.

Every generated document is clearly marked as synthetic in its header.

Usage:
    python scripts/make_fixture_pack.py [--out PATH]
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = REPO_ROOT / "fixtures" / "synthetic_datapack"

SNAPSHOT_UTC = "2026-08-21T23:59:00Z"


# --------------------------------------------------------------------------
# PDF helpers
# --------------------------------------------------------------------------
def render_pdf(path: Path, title: str, meta_lines: list[str], sections: list[tuple[int, str, str]]) -> None:
    pdf = FPDF()
    pdf.add_page()

    def para(text: str, height: float) -> None:
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(pdf.epw, height, text)
        pdf.set_x(pdf.l_margin)

    pdf.set_font("Helvetica", "B", 15)
    para(title, 8)
    pdf.ln(1)
    pdf.set_font("Helvetica", "", 10)
    for line in meta_lines:
        para(line, 5)
    pdf.ln(3)
    for level, heading, body in sections:
        pdf.set_font("Helvetica", "B", 12 if level == 1 else 10.5)
        para(heading, 7)
        pdf.set_font("Helvetica", "", 10)
        for para_text in [p.strip() for p in body.strip().split("\n\n") if p.strip()]:
            para(" ".join(para_text.split()), 5)
            pdf.ln(1)
        pdf.ln(1.5)
    path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(path))


# --------------------------------------------------------------------------
# Document contents (all ASCII so core PDF fonts stay happy)
# --------------------------------------------------------------------------
DOC01_SECTIONS = [
    (1, "1. Scope and Definitions",
     "This policy applies to all ParcelPilot business customers on Standard, Growth and "
     "Enterprise plans unless a signed customer agreement states otherwise.\n\n"
     "'Pickup commences' means the moment a driver arrives at the collection address and "
     "starts loading. 'Business hours' are 08:00-20:00 UTC Monday to Friday."),
    (1, "2. Support Channels and Hours",
     "The support portal is available 24/7 at portal.parcelpilot.example.\n\n"
     "Email support is monitored during business hours. Phone support is available to "
     "Growth and Enterprise customers during business hours."),
    (1, "3. Severity Levels and Response SLAs",
     "P1 Critical - platform unusable or cold-chain failure in progress. First response "
     "within 1 hour, target resolution within 8 business hours.\n\n"
     "P2 Major - major feature degraded, no workaround. First response within 4 business "
     "hours, target resolution within 24 business hours.\n\n"
     "P3 Minor - minor issue or question. First response within 1 business day, target "
     "resolution within 3 business days."),
    (1, "4. Cancellation Policy",
     "A booking may be cancelled free of charge if the cancellation is requested before "
     "pickup commences AND within 60 minutes of the booking being placed.\n\n"
     "If pickup has already commenced, cancellation incurs a fee equal to 25% of the "
     "order value with a minimum fee of USD 40. Prepaid amounts are refunded minus the "
     "applicable fee within 10 business days."),
    (1, "5. Late Pickup Compensation",
     "If a pickup occurs more than 90 minutes after the scheduled pickup time, the "
     "customer receives a service credit of USD 25.\n\n"
     "If the delay exceeds 180 minutes, the credit is 10% of the order value, capped at "
     "USD 150.\n\n"
     "Compensation must be requested within 14 days of the incident."),
    (1, "6. Service Credit Administration",
     "Service credits are applied to future invoices, are not redeemable for cash, and "
     "are capped at USD 500 per account per calendar month.\n\n"
     "Any single credit above USD 250 requires approval by a ParcelPilot operations "
     "manager before it is issued."),
    (1, "7. Escalation Path",
     "Escalate to a human support manager when policy is unclear, when sources conflict "
     "and cannot be resolved, when a request falls outside this policy, or when the "
     "customer requests a human. Escalations must reference the affected order or "
     "ticket identifiers."),
]

DOC02_SECTIONS = [
    (1, "1. Scope",
     "This version applies until superseded. It is retained for auditing historical "
     "tickets raised before the effective date of version 3."),
    (1, "2. Severity Levels and Response SLAs",
     "P1 first response within 2 business hours. P2 first response within 8 business "
     "hours. P3 first response within 2 business days."),
    (1, "3. Cancellation Policy",
     "Cancellations are free if requested within 120 minutes of booking and before "
     "pickup commences.\n\n"
     "After pickup commences the cancellation fee is 15% of order value, minimum USD 25."),
    (1, "4. Late Pickup Compensation",
     "Pickups delayed more than 120 minutes receive a USD 15 service credit. Delays "
     "over 240 minutes receive a flat USD 50 credit."),
]

DOC03_SECTIONS = [
    (1, "1. Purpose",
     "Standard operating procedure for ParcelPilot support staff handling cancellations "
     "and service-credit requests. Follows the Customer Support Policy; where the "
     "customer has a signed agreement, the agreement takes precedence."),
    (1, "2. Identity and Account Verification",
     "Confirm the requester acts for the account before disclosing any order, invoice or "
     "ticket detail. Never act on or disclose data belonging to a different account."),
    (1, "3. Data to Collect Before Deciding",
     "Retrieve the order record and confirm scheduled versus actual timestamps. Confirm "
     "the request was raised within the applicable claim window. Check account standing "
     "(no unpaid invoices older than 30 days). Identify whether an account-specific "
     "agreement exists and read its terms."),
    (1, "4. Cancellation Handling Procedure",
     "Step 1: retrieve booking and confirm status. Step 2: determine whether pickup had "
     "commenced. Step 3: compute the fee under the governing source - customer agreement "
     "first, otherwise the current support policy. Step 4: quote the fee, obtain explicit "
     "confirmation, then record the cancellation. Step 5: log the action against the "
     "ticket."),
    (1, "5. Service Credit Eligibility Checklist",
     "All four conditions must hold unless the customer agreement provides otherwise:\n\n"
     "a) A breach is confirmed from the order record timestamps themselves.\n\n"
     "b) The request was raised within 14 days of the incident (or the window stated in "
     "the customer agreement).\n\n"
     "c) The account is in good standing, meaning no unpaid invoices older than 30 "
     "days.\n\n"
     "d) The amount fits the monthly cap and approval matrix below."),
    (1, "6. Approval Matrix",
     "Credits up to USD 250 may be approved by the handling agent. Above USD 250 an "
     "operations manager must approve. Above USD 500 a director must approve."),
    (1, "7. Use of Historical Tickets",
     "Past ticket outcomes and notes are context only. They show how similar situations "
     "were handled but are never an authoritative source for fees, entitlements or "
     "eligibility. Always cite the governing policy or agreement instead."),
]

DOC04_SECTIONS = [
    (1, "1. Platform Overview",
     "ParcelPilot provides booking, dispatch, live tracking and proof-of-delivery for "
     "business logistics customers."),
    (1, "2. Tracking and ETA Behaviour",
     "Driver devices report GPS position every 30 seconds in urban areas and every 2 "
     "minutes on rural routes. Estimated arrival times are recalculated on each ping."),
    (1, "3. Webhooks and Integrations",
     "Events include order.booked, pickup.completed and delivery.completed. Consumers "
     "must deduplicate on the event_id field; retries may occasionally deliver an event "
     "more than once."),
    (1, "4. Known Issues Register",
     "KB-2025-011 (OPEN): ETA jitter on rural routes caused by GPS drift. ETAs may swing "
     "by up to 45 minutes between updates. Workaround: refresh the tracking view or use "
     "the dispatch console for a stable ETA. Affects Standard, Express and Cold Chain "
     "services on rural legs.\n\n"
     "KB-2025-007 (MONITORING): Duplicate webhook deliveries under retry storms after "
     "platform incidents. Consumers must deduplicate on event_id.\n\n"
     "KB-2024-102 (RESOLVED in v2.9.1): Cold-chain sensor lag caused delayed temperature "
     "alerts."),
]

DOC05_SECTIONS = [
    (1, "1. Parties and Term",
     "This Enterprise Agreement ('Agreement') is between ParcelPilot Inc. and Northstar "
     "Logistics (Account ID ACC-001), effective 2026-03-01."),
    (1, "2. Order of Precedence",
     "Where this Agreement conflicts with the standard ParcelPilot Customer Support "
     "Policy, this Agreement prevails for Northstar Logistics."),
    (1, "3. Cancellation Terms",
     "Northstar may cancel any booking free of charge at any time before pickup "
     "commences.\n\n"
     "Once pickup has commenced, cancellation incurs a flat administrative fee of USD 75 "
     "per order regardless of order value."),
    (1, "4. Service Levels",
     "P1 first response within 15 minutes, 24x7x365. Monthly platform uptime commitment "
     "99.9%."),
    (1, "5. Late Delivery Compensation",
     "If a delivery completes more than 4 hours after the promised delivery time, "
     "Northstar receives a service credit of 5% of its monthly recurring platform fee "
     "per occurrence, capped at 20% of the monthly fee per calendar month.\n\n"
     "Claims must be raised within 30 days of the late delivery."),
    (1, "6. Support Team",
     "Northstar is served by a dedicated customer success manager reachable at "
     "northstar-csm@parcelpilot.example."),
]

DOC06_SECTIONS = [
    (1, "1. Parties and Term",
     "This Service Agreement is between ParcelPilot Inc. and LumenWorks Ltd (Account ID "
     "ACC-002), effective 2026-01-15."),
    (1, "2. Order of Precedence",
     "Where this Agreement conflicts with the standard ParcelPilot Customer Support "
     "Policy, this Agreement prevails for LumenWorks Ltd."),
    (1, "3. Cancellation Terms",
     "Bookings may be cancelled free of charge before pickup commences.\n\n"
     "After pickup commences, the cancellation fee is the lesser of USD 100 or 20% of "
     "the order value."),
    (1, "4. Pickup Reliability Credit",
     "If a pickup occurs more than 60 minutes after the scheduled pickup time, "
     "LumenWorks receives a USD 50 service credit per occurrence. Claims must be raised "
     "within 14 days."),
    (1, "5. Support Hours",
     "Extended support hours 06:00-22:00 CET Monday to Saturday. P1 first response "
     "within 30 minutes."),
]


def build_pdfs(out: Path) -> None:
    render_pdf(
        out / "01_Support_Policy_v3_CURRENT.pdf",
        "ParcelPilot Customer Support Policy",
        ["Document ID: PP-POL-001", "Version: 3", "Status: CURRENT",
         "Effective date: 2026-05-01", "[SYNTHETIC FIXTURE - not the real assessment pack]"],
        DOC01_SECTIONS,
    )
    render_pdf(
        out / "02_Support_Policy_v2_DEPRECATED.pdf",
        "ParcelPilot Customer Support Policy",
        ["Document ID: PP-POL-001", "Version: 2", "Status: DEPRECATED",
         "Superseded by version 3 effective 2026-05-01",
         "[SYNTHETIC FIXTURE - not the real assessment pack]"],
        DOC02_SECTIONS,
    )
    render_pdf(
        out / "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
        "Cancellation and Service Credit SOP",
        ["Document ID: PP-SOP-004", "Version: 4", "Status: CURRENT",
         "[SYNTHETIC FIXTURE - not the real assessment pack]"],
        DOC03_SECTIONS,
    )
    render_pdf(
        out / "04_Product_Operations_Guide_and_Known_Issues.pdf",
        "Product Operations Guide and Known Issues",
        ["Document ID: PP-OPS-002", "Version: 2", "Status: CURRENT",
         "[SYNTHETIC FIXTURE - not the real assessment pack]"],
        DOC04_SECTIONS,
    )
    render_pdf(
        out / "05_Northstar_Logistics_Enterprise_Agreement.pdf",
        "Northstar Logistics Enterprise Agreement",
        ["Document ID: PP-AGR-NS-001", "Version: 1", "Status: CURRENT",
         "Effective date: 2026-03-01", "[SYNTHETIC FIXTURE - not the real assessment pack]"],
        DOC05_SECTIONS,
    )
    render_pdf(
        out / "06_LumenWorks_Service_Agreement.pdf",
        "LumenWorks Ltd Service Agreement",
        ["Document ID: PP-AGR-LW-001", "Version: 1", "Status: CURRENT",
         "Effective date: 2026-01-15", "[SYNTHETIC FIXTURE - not the real assessment pack]"],
        DOC06_SECTIONS,
    )


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------
ACCOUNTS_COLS = ["account_id", "account_name", "tier", "primary_contact", "good_standing", "onboarded_at"]
ORDER_COLS = ["order_id", "account_id", "service_type", "booked_at", "scheduled_pickup_at",
              "actual_pickup_at", "promised_delivery_at", "delivered_at", "status", "order_value_usd"]
TICKET_COLS = ["ticket_id", "account_id", "order_id", "category", "subject", "description",
               "priority", "created_at", "first_response_at", "resolved_at", "status", "resolution_note"]

ACCOUNTS = [
    ("ACC-001", "Northstar Logistics", "Enterprise", "ops@northstar.example", 1, "2024-11-02T09:00:00Z"),
    ("ACC-002", "LumenWorks Ltd", "Enterprise", "support@lumenworks.example", 1, "2025-02-17T09:00:00Z"),
    ("ACC-003", "BrightCart Commerce", "Growth", "logistics@brightcart.example", 1, "2025-06-08T09:00:00Z"),
    ("ACC-004", "FreshFleet Grocers", "Starter", "admin@freshfleet.example", 0, "2025-09-23T09:00:00Z"),
    ("ACC-005", "SwiftMed Supplies", "Growth", "dispatch@swiftmed.example", 1, "2025-04-30T09:00:00Z"),
]

# (order_id, acct, svc, booked, sched_pickup, actual_pickup, promised_del, delivered, status, value)
ORDERS = [
    ("ORD-1001", "ACC-001", "Express", "2026-08-17T14:05:00Z", "2026-08-18T09:00:00Z", "2026-08-18T09:47:00Z", "2026-08-19T17:00:00Z", "2026-08-20T02:30:00Z", "delivered", 840.0),
    ("ORD-1002", "ACC-002", "Same-Day", "2026-08-12T10:12:00Z", "2026-08-12T11:00:00Z", "2026-08-12T12:02:00Z", "2026-08-12T16:00:00Z", "2026-08-12T15:40:00Z", "delivered", 210.0),
    ("ORD-1003", "ACC-005", "Express", "2026-08-19T06:40:00Z", "2026-08-19T08:00:00Z", "2026-08-19T09:35:00Z", "2026-08-19T20:00:00Z", "2026-08-19T21:10:00Z", "delivered", 430.0),
    ("ORD-1004", "ACC-003", "Standard", "2026-08-03T11:00:00Z", "2026-08-03T15:00:00Z", "2026-08-03T15:04:00Z", "2026-08-04T12:00:00Z", "2026-08-04T11:22:00Z", "delivered", 150.0),
    ("ORD-1005", "ACC-004", "Cold Chain", "2026-08-06T07:15:00Z", "2026-08-06T09:30:00Z", None, "2026-08-07T09:00:00Z", None, "cancelled_before_pickup", 95.0),
    ("ORD-1006", "ACC-003", "Standard", "2026-08-08T08:00:00Z", "2026-08-08T12:00:00Z", "2026-08-08T12:05:00Z", "2026-08-09T12:00:00Z", None, "cancelled_after_pickup", 320.0),
    ("ORD-1007", "ACC-002", "Standard", "2026-08-10T05:55:00Z", "2026-08-10T07:00:00Z", "2026-08-10T07:02:00Z", "2026-08-10T19:00:00Z", "2026-08-10T18:31:00Z", "delivered", 260.0),
    ("ORD-1008", "ACC-005", "Standard", "2026-08-10T09:20:00Z", "2026-08-10T13:00:00Z", "2026-08-10T13:03:00Z", "2026-08-11T18:00:00Z", "2026-08-11T17:44:00Z", "delivered", 175.0),
    ("ORD-1009", "ACC-001", "Standard", "2026-08-11T08:10:00Z", "2026-08-11T12:30:00Z", "2026-08-11T12:33:00Z", "2026-08-12T17:00:00Z", "2026-08-12T16:52:00Z", "delivered", 610.0),
    ("ORD-1010", "ACC-003", "Express", "2026-08-11T13:45:00Z", "2026-08-11T15:30:00Z", "2026-08-11T15:29:00Z", "2026-08-11T22:00:00Z", "2026-08-11T21:38:00Z", "delivered", 385.0),
    ("ORD-1011", "ACC-004", "Standard", "2026-08-12T10:05:00Z", "2026-08-12T14:00:00Z", "2026-08-12T14:02:00Z", "2026-08-13T12:00:00Z", "2026-08-13T11:49:00Z", "delivered", 140.0),
    ("ORD-1012", "ACC-002", "Same-Day", "2026-08-13T07:30:00Z", "2026-08-13T08:30:00Z", "2026-08-13T08:33:00Z", "2026-08-13T14:00:00Z", "2026-08-13T13:47:00Z", "delivered", 190.0),
    ("ORD-1013", "ACC-005", "Standard", "2026-08-13T11:25:00Z", "2026-08-13T15:00:00Z", "2026-08-13T15:04:00Z", "2026-08-14T12:00:00Z", "2026-08-14T11:58:00Z", "delivered", 205.0),
    ("ORD-1014", "ACC-001", "Same-Day", "2026-08-14T09:00:00Z", "2026-08-14T10:00:00Z", "2026-08-14T10:48:00Z", "2026-08-14T16:00:00Z", "2026-08-14T15:30:00Z", "delivered", 275.0),
    ("ORD-1015", "ACC-003", "Standard", "2026-08-15T12:00:00Z", "2026-08-15T16:00:00Z", "2026-08-15T16:02:00Z", "2026-08-16T12:00:00Z", "2026-08-16T11:40:00Z", "delivered", 160.0),
    ("ORD-1016", "ACC-004", "Express", "2026-08-16T06:50:00Z", "2026-08-16T08:30:00Z", "2026-08-16T08:34:00Z", "2026-08-16T20:00:00Z", "2026-08-16T19:22:00Z", "delivered", 350.0),
    ("ORD-1017", "ACC-002", "Cold Chain", "2026-08-17T05:40:00Z", "2026-08-17T07:15:00Z", "2026-08-17T07:20:00Z", "2026-08-17T15:00:00Z", "2026-08-17T14:41:00Z", "delivered", 520.0),
    ("ORD-1018", "ACC-005", "Same-Day", "2026-08-18T08:20:00Z", "2026-08-18T09:20:00Z", "2026-08-18T09:22:00Z", "2026-08-18T15:00:00Z", "2026-08-18T14:35:00Z", "delivered", 230.0),
    ("ORD-1019", "ACC-001", "Standard", "2026-08-18T13:10:00Z", "2026-08-18T17:00:00Z", "2026-08-18T17:03:00Z", "2026-08-19T12:00:00Z", "2026-08-19T11:39:00Z", "delivered", 480.0),
    ("ORD-1020", "ACC-003", "Standard", "2026-08-19T09:45:00Z", "2026-08-19T13:30:00Z", "2026-08-19T13:33:00Z", "2026-08-20T12:00:00Z", "2026-08-20T11:51:00Z", "delivered", 185.0),
    ("ORD-1021", "ACC-002", "Express", "2026-08-19T14:30:00Z", "2026-08-19T16:15:00Z", "2026-08-19T16:18:00Z", "2026-08-19T23:30:00Z", "2026-08-19T23:05:00Z", "delivered", 640.0),
    ("ORD-1022", "ACC-004", "Standard", "2026-08-20T07:05:00Z", "2026-08-20T11:00:00Z", "2026-08-20T11:03:00Z", "2026-08-21T12:00:00Z", "2026-08-21T11:47:00Z", "delivered", 125.0),
    ("ORD-1023", "ACC-005", "Cold Chain", "2026-08-20T10:40:00Z", "2026-08-20T12:30:00Z", "2026-08-20T12:36:00Z", "2026-08-20T20:30:00Z", "2026-08-20T20:04:00Z", "delivered", 710.0),
    ("ORD-1024", "ACC-001", "Express", "2026-08-21T05:20:00Z", "2026-08-21T07:00:00Z", "2026-08-21T07:04:00Z", "2026-08-21T18:00:00Z", None, "in_transit", 900.0),
    ("ORD-1025", "ACC-003", "Same-Day", "2026-08-21T11:35:00Z", "2026-08-21T12:30:00Z", "2026-08-21T12:33:00Z", "2026-08-21T19:00:00Z", None, "in_transit", 240.0),
    ("ORD-1026", "ACC-002", "Standard", "2026-08-04T09:10:00Z", "2026-08-04T11:00:00Z", "2026-08-04T13:35:00Z", "2026-08-05T12:00:00Z", "2026-08-05T11:30:00Z", "delivered", 300.0),
]

# (ticket_id, acct, order, cat, subject, desc, prio, created, first_resp, resolved, status, note)
TICKETS = [
    ("TCK-2001", "ACC-003", "ORD-1010", "tracking", "ETA keeps jumping by 40 minutes",
     "Rural leg ETA swings wildly between refreshes.", "P3",
     "2026-08-11T18:20:00Z", "2026-08-11T19:02:00Z", "2026-08-12T09:10:00Z", "resolved",
     "Explained known issue KB-2025-011; suggested dispatch console ETA."),
    ("TCK-2002", "ACC-005", "ORD-1008", "tracking", "Tracking ETA unreliable on rural route",
     "Customer reports ETA moving by nearly an hour.", "P3",
     "2026-08-11T10:44:00Z", "2026-08-11T11:31:00Z", "2026-08-12T08:47:00Z", "resolved",
     "Known issue KB-2025-011; workaround provided."),
    ("TCK-2003", "ACC-004", "ORD-1011", "tracking", "ETA jumped backwards twice",
     "ETA decreased after departure which confused our warehouse.", "P3",
     "2026-08-12T15:12:00Z", "2026-08-12T16:05:00Z", "2026-08-13T10:02:00Z", "resolved",
     "Known issue KB-2025-011."),
    ("TCK-2004", "ACC-002", "ORD-1012", "tracking", "Delivery ETA unstable all afternoon",
     "Same-day delivery ETA kept changing by 30-45 minutes.", "P3",
     "2026-08-13T12:41:00Z", "2026-08-13T13:20:00Z", "2026-08-14T09:15:00Z", "resolved",
     "Known issue KB-2025-011; monitoring."),
    ("TCK-2005", "ACC-003", "ORD-1015", "tracking", "Another rural ETA jump",
     "Second week in a row with unstable ETAs on this lane.", "P3",
     "2026-08-15T17:30:00Z", "2026-08-15T18:10:00Z", "2026-08-16T08:55:00Z", "resolved",
     "Known issue KB-2025-011; escalated to engineering for prioritisation."),
    ("TCK-2006", "ACC-005", "ORD-1013", "tracking", "ETA wrong by almost an hour",
     "Dispatch had to call the driver manually.", "P3",
     "2026-08-13T16:02:00Z", "2026-08-13T16:44:00Z", "2026-08-14T11:20:00Z", "resolved",
     "Known issue KB-2025-011."),
    ("TCK-2007", "ACC-001", "ORD-1001", "billing", "Late delivery compensation request ORD-1001",
     "Promised delivery 2026-08-19 17:00 UTC, delivered 2026-08-20 02:30 UTC. Requesting compensation.",
     "P2", "2026-08-20T10:15:00Z", None, None, "open", ""),
    ("TCK-2008", "ACC-002", "ORD-1002", "billing", "Credit request for delayed pickup ORD-1002",
     "Pickup was more than an hour late. Requesting service credit under our agreement.",
     "P3", "2026-09-01T09:00:00Z", None, None, "open", ""),
    ("TCK-2009", "ACC-004", "ORD-1005", "cancellation", "Cancellation fee question ORD-1005",
     "Order was cancelled before any driver arrived. Asking why a fee might apply.", "P2",
     "2026-08-06T11:05:00Z", "2026-08-06T11:40:00Z", None, "in_progress", ""),
    ("TCK-2010", "ACC-003", "ORD-1006", "cancellation", "Fee charged for cancelled order ORD-1006",
     "Cancelled mid-delivery; querying the 25 percent fee.", "P3",
     "2026-08-08T14:22:00Z", "2026-08-08T15:01:00Z", "2026-08-09T09:30:00Z", "resolved",
     "Pickup had commenced; standard policy v3 fee applied (25% of 320 = USD 80)."),
    ("TCK-2011", "ACC-001", "ORD-1014", "sla", "Late same-day pickup ORD-1014",
     "Driver arrived 48 minutes after window; asking if this breaches SLA.", "P3",
     "2026-08-14T16:45:00Z", "2026-08-14T17:20:00Z", "2026-08-15T10:05:00Z", "resolved",
     "Below the 90 minute standard threshold; no credit due."),
    ("TCK-2012", "ACC-005", "ORD-1003", "sla", "Pickup over two hours late ORD-1003",
     "Scheduled 08:00, collected 09:35. Requesting applicable credit.", "P2",
     "2026-08-19T12:30:00Z", "2026-08-19T13:02:00Z", None, "in_progress", ""),
    ("TCK-2013", "ACC-005", "ORD-1023", "damaged", "Cold chain temp excursion alert ORD-1023",
     "Sensor logged 9.1C for 25 minutes mid-route. Product quarantined.", "P2",
     "2026-08-21T08:12:00Z", "2026-08-21T08:47:00Z", None, "open", ""),
    ("TCK-2014", "ACC-002", "ORD-1026", "sla", "Pickup 155 minutes late ORD-1026",
     "Scheduled 11:00, actual 13:35 on 2026-08-04.", "P3",
     "2026-08-05T10:20:00Z", "2026-08-05T11:02:00Z", "2026-08-06T09:40:00Z", "resolved",
     "USD 50 pickup reliability credit issued under the LumenWorks agreement."),
    ("TCK-2015", "ACC-001", "ORD-1009", "webhook", "Duplicate webhook events received",
     "Integration received delivery.completed twice for the same order.", "P3",
     "2026-08-12T09:15:00Z", "2026-08-12T09:58:00Z", "2026-08-13T08:30:00Z", "resolved",
     "Known issue KB-2025-007; advised dedupe on event_id."),
    ("TCK-2016", "ACC-003", "ORD-1004", "invoice", "Invoice line does not match quote",
     "August invoice shows USD 155 vs quoted USD 150.", "P3",
     "2026-08-25T09:40:00Z", "2026-08-25T10:22:00Z", "2026-08-26T14:10:00Z", "resolved",
     "Fuel surcharge added post-quote; credited the difference."),
    ("TCK-2017", "ACC-004", "ORD-1016", "tracking", "Express delivered 38 minutes late",
     "Promise missed on express lane; asking about remedies.", "P3",
     "2026-08-16T20:05:00Z", "2026-08-16T20:41:00Z", "2026-08-17T09:25:00Z", "resolved",
     "No credit entitlement under starter plan terms; apology issued."),
    ("TCK-2018", "ACC-001", "ORD-1024", "tracking", "Where is my urgent shipment?",
     "In-transit since 07:04; customer asks for proactive updates.", "P2",
     "2026-08-21T13:30:00Z", "2026-08-21T13:44:00Z", None, "in_progress", ""),
    ("TCK-2019", "ACC-002", "ORD-1017", "damaged", "Condensation inside cold chain parcel",
     "Packaging damp on arrival; goods unaffected but packaging questioned.", "P3",
     "2026-08-17T16:10:00Z", "2026-08-17T16:48:00Z", "2026-08-18T10:12:00Z", "resolved",
     "Advised packaging condensation is cosmetic in humid conditions; replaced liner stock."),
    ("TCK-2020", "ACC-005", "ORD-1018", "cancellation", "Cancel remaining leg ORD-1018",
     "Recipient closed early; requested cancellation after pickup started.", "P2",
     "2026-08-18T11:25:00Z", "2026-08-18T11:59:00Z", "2026-08-18T13:40:00Z", "resolved",
     "Delivered anyway before cancellation processed; no fee charged."),
    ("TCK-2021", "ACC-003", "ORD-1020", "sla", "Will ORD-1020 hit its promise time?",
     "Customer watching a critical restock.", "P3",
     "2026-08-20T08:00:00Z", "2026-08-20T08:35:00Z", "2026-08-20T11:51:00Z", "resolved",
     "Delivered on time 2026-08-20 11:51 UTC."),
    ("TCK-2022", "ACC-004", "ORD-1022", "invoice", "Requesting credit for late delivery ORD-1022",
     "Delivered 2026-08-21 11:47 vs promise 12:00 - actually early?", "P3",
     "2026-08-21T15:00:00Z", None, None, "open", ""),
]


def build_xlsx(path: Path) -> None:
    wb = Workbook()

    readme = wb.active
    readme.title = "README"
    for row in [
        ("dataset_name", "ParcelPilot Assessment Data"),
        ("dataset_version", "1.4"),
        ("dataset_snapshot_utc", SNAPSHOT_UTC),
        ("prepared_by", "data-team@parcelpilot.example"),
        ("sheet_accounts", "One row per customer account. good_standing=0 means unpaid invoices older than 30 days."),
        ("sheet_orders", "One row per order. All timestamps UTC ISO-8601. Null actual_pickup_at means pickup never happened."),
        ("sheet_tickets", "One row per support ticket. resolution_note is context only, never an authoritative source."),
        ("notes", "Synthetic fixture dataset generated by scripts/make_fixture_pack.py - stands in for the real assessment pack."),
    ]:
        readme.append(row)

    ws = wb.create_sheet("accounts")
    ws.append(ACCOUNTS_COLS)
    for r in ACCOUNTS:
        ws.append(list(r))

    wo = wb.create_sheet("orders")
    wo.append(ORDER_COLS)
    for r in ORDERS:
        wo.append([None if v is None else v for v in r])

    wt = wb.create_sheet("tickets")
    wt.append(TICKET_COLS)
    for r in TICKETS:
        wt.append(["" if v == "" and c == "resolution_note" else v for c, v in zip(TICKET_COLS, r)])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    out = args.out.resolve()
    build_pdfs(out)
    build_xlsx(out / "ParcelPilot_Assessment_Data.xlsx")
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    print(f"[make_fixture_pack] wrote synthetic pack to {out} at {stamp}")
    print("[make_fixture_pack] NOTE: these files are SYNTHETIC fixtures, not the real assessment pack.")


if __name__ == "__main__":
    main()
